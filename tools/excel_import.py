#!/bin/python

from openpyxl import *
from datetime import timedelta, datetime, time
import sqlite3
import sys

# constant
DB_FILE = "../strengthy.db"
USER_ID = 1
LEGDAY_WID = 1
PUSHDAY_WID = 3
PULLDAY_WID = 4


def fix_name(name):
    name = name.strip()
    if name == "Rope Tricep Pulldown":
        return "Cable Tricep Pushdown"
    elif name == "Cable Row":
        return "Chest Row"
    elif name == "Dumbell Shrugs":
        return "Dumbbell Shrugs"
    elif name == "Roller":
        return "Ab Roller"
    elif name == "Lateral Raise":
        return "Delt Raise"
    else:
        return name


def parse_workout(workout_id, rows):
    # Iterate over exercises
    sets = []
    for row in rows:
        # Only get rows that have an exercise
        name = sheet.cell(row, 1).value
        if not name:
            continue
        name = fix_name(name)

        # Iterate over sets
        for col in range(2, 7):
            lbs = sheet.cell(row, col).value
            reps = sheet.cell(row + 1, col).value

            if (
                lbs
                and reps
                and not isinstance(reps, time)
                and not isinstance(lbs, time)
            ):
                sets.append((name, float(lbs), int(reps)))

    global date
    workout = {"workout_id": workout_id, "sets": sets, "finished": date}
    # Update date
    date -= timedelta(days=1)

    return workout


exercise_ids = {}


def insert_workout(cur, workout):
    cur.execute(
        "INSERT INTO workout_records(finished, user_id, workout_id) VALUES (?, ?, ?)",
        (
            workout["finished"],
            USER_ID,
            int(workout["workout_id"]),
        ),
    )
    workout_record_id = cur.lastrowid

    for s in workout["sets"]:
        # Get exercise id if not cached
        if s[0] not in exercise_ids:
            row = cur.execute(
                "SELECT * FROM exercises WHERE name = :name", {"name": s[0]}
            ).fetchone()

            # Invalid exercise
            if not row:
                print("Could not find id for", s[0])
                continue

            exercise_ids[s[0]] = int(row[0])

        exercise_id = exercise_ids[s[0]]

        cur.execute(
            "INSERT INTO set_records(workout_record_id, lbs, reps, exercise_id) VALUES (?, ?, ?, ?)",
            (workout_record_id, s[1], s[2], exercise_id),
        )


# Load
print("Importing data...")
workbook = load_workbook(sys.argv[1])

# Get all "week sheets"
print("Parsing sheets...")
week_sheets = []
for sheet in workbook:
    if sheet.title.startswith("Week "):
        week_sheets.append(sheet)

# Parse all workouts
print("Parsing workouts...")

workouts = []
date = datetime.today()
date = date.replace(hour=15, minute=0, second=0)

for sheet in week_sheets:
    print(f"Parsing {sheet.title}...")
    workouts.append(parse_workout(LEGDAY_WID, range(2, 22)))
    workouts.append(parse_workout(PUSHDAY_WID, range(25, 39)))
    workouts.append(parse_workout(PULLDAY_WID, range(42, 56)))
    workouts.append(parse_workout(LEGDAY_WID, range(59, 79)))
    workouts.append(parse_workout(PUSHDAY_WID, range(82, 96)))
    workouts.append(parse_workout(PULLDAY_WID, range(99, 113)))
    # Skip sunday
    date -= timedelta(days=1)

# Load database
print("Connecting to database...")
conn = sqlite3.connect(DB_FILE)
cur = conn.cursor()

# Insert workouts into the database
for workout in workouts:
    insert_workout(cur, workout)

conn.commit()
conn.close()
